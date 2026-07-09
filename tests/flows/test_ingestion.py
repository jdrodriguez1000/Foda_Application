"""Tests unitarios de Ingestion (feature ingestion, banda tracer_bullet).

Fuente: 600_features/ingestion/tracer_bullet/spec.md (CA-xx) y plan.md
(casos TDD, TSK-14..TSK-36). Bucle TDD: un test por caso, ejecutado en orden
(state.json -> stages.tdd.cases). Este archivo arranca con el caso 1
(CA-14): fixture minimo (un unico dataset "ventas"/"ventas.csv" coma,
DS-ING-7/DS-ING-8) con contract_data.json (fuente de los archivos esperados)
y map_client_data.json (fuente de las columnas esperadas) coherentes entre
si, mas el archivo crudo bajo el landing.
"""

import io
import json
from pathlib import Path

import openpyxl
import pytest

from foda.core.context import ClientContext
from foda.core.flow import Artifact, Flow, FlowContractError
from foda.core.scaffold import create_client
from foda.flows.f030_ingestion.ingestion import Ingestion

_VENTAS_HEADER = "fecha,sede,clase,cantidad,precio_unitario"
_VENTAS_ROWS = [
    "2024-01-01,Sede Centro,Agua 600ml,10,1200",
    "2024-01-02,Sede Norte,Cola 1.5L,5,2500",
    "2024-01-03,Sede Centro,Papas 45g,20,900",
]


def _contract_data_minimo() -> dict:
    """DS-ING-8: fuente de los archivos esperados. Un unico dataset "ventas"
    con un unico archivo "ventas.csv" (caso 1, subconjunto minimo del fixture
    DS-ING-7)."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "historical_data": {
            "datasets": [
                {
                    "kind": "ventas",
                    "source_medium": "csv",
                    "periodicity": "mensual",
                    "files": [
                        {
                            "name": "ventas.csv",
                            "period_start": "2023-01-01",
                            "period_end": "2025-12-31",
                        }
                    ],
                }
            ]
        },
    }


def _map_client_data_minimo() -> dict:
    """DS-ING-8: fuente de las columnas esperadas por dataset (fields[] con
    name/required), emparejadas por kind con el dataset homologo del
    contrato. Coherente con _contract_data_minimo (mismo kind "ventas")."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "datasets": [
            {
                "kind": "ventas",
                "fields": [
                    {"name": "fecha", "required": True},
                    {"name": "sede", "required": True},
                    {"name": "clase", "required": True},
                    {"name": "cantidad", "required": True},
                    {"name": "precio_unitario", "required": False},
                ],
            }
        ],
    }


def _build_ctx(
    tmp_path: Path,
    contract_data: dict,
    map_client_data: dict,
    files: dict[str, str | bytes],
) -> ClientContext:
    """DS-ING-7/DS-ING-8: construye un ClientContext bajo tmp_path con
    contract_data.json + map_client_data.json bajo ctx.outputs_dir, y uno o
    mas archivos crudos (name -> contenido, texto delimitado o bytes
    binarios p. ej. .xlsx) bajo ctx.inputs_dir/"030_ingestion". Helper
    compartido por las fixtures de cada caso del bucle TDD (evita duplicar
    el andamiaje de directorios)."""
    clients_root = tmp_path / "clients"
    create_client("ABC", clients_root)
    ctx = ClientContext("ABC", clients_root)

    contrato_path = ctx.outputs_dir / "010_discovery/contract_data.json"
    contrato_path.parent.mkdir(parents=True)
    contrato_path.write_text(
        json.dumps(contract_data, ensure_ascii=False), encoding="utf-8"
    )

    mapa_path = ctx.outputs_dir / "020_onboarding/map_client_data.json"
    mapa_path.parent.mkdir(parents=True)
    mapa_path.write_text(
        json.dumps(map_client_data, ensure_ascii=False), encoding="utf-8"
    )

    landing_dir = ctx.inputs_dir / "030_ingestion"
    landing_dir.mkdir(parents=True)
    for name, content in files.items():
        if isinstance(content, bytes):
            (landing_dir / name).write_bytes(content)
        else:
            (landing_dir / name).write_text(content, encoding="utf-8")

    return ctx


def _build_ctx_fixture_minimo(tmp_path: Path) -> ClientContext:
    """DS-ING-7 (subconjunto minimo, casos 1-3): ClientContext con
    contract_data.json + map_client_data.json coherentes entre si (dataset
    "ventas" unico) y el archivo crudo ventas.csv (separador coma)."""
    return _build_ctx(
        tmp_path,
        _contract_data_minimo(),
        _map_client_data_minimo(),
        {"ventas.csv": "\n".join([_VENTAS_HEADER, *_VENTAS_ROWS]) + "\n"},
    )


def test_run_sobre_fixture_minimo_escribe_ingestion_report_y_lo_incluye_en_outputs(
    tmp_path: Path,
) -> None:
    """Caso 1 (CA-14): run(ctx) sobre el fixture minimo (dataset ventas /
    ventas.csv coma) escribe ingestion_report.json en
    ctx.outputs_dir / "030_ingestion/ingestion_report.json" y lo incluye en
    FlowResult.outputs; el archivo existe en disco."""
    ctx = _build_ctx_fixture_minimo(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_esperada = ctx.outputs_dir / "030_ingestion/ingestion_report.json"

    assert ruta_esperada.exists()
    assert ruta_esperada in result.outputs


def test_ingestion_hereda_flow_requires_produces_y_4_fases_sin_sobreescribir_run() -> None:
    """Caso 2 (CA-20): Ingestion hereda de Flow, declara requires/produces con
    los Artifact exactos de la spec y completa las 4 fases del template
    method (load_inputs, validate, execute, write_outputs) sin sobreescribir
    run. El caso 1 ya cubre que run() (heredado) produce el reporte
    end-to-end; este caso verifica la estructura declarativa de la clase que
    ese caso no comprueba: que Ingestion sobreescribe explicitamente las 4
    fases (incluida validate, aunque delegue en super().validate) y no
    sobreescribe run."""
    assert issubclass(Ingestion, Flow)

    assert Ingestion.requires == [
        Artifact(
            name="contract_data",
            base="outputs",
            relative="010_discovery/contract_data.json",
        ),
        Artifact(
            name="map_client_data",
            base="outputs",
            relative="020_onboarding/map_client_data.json",
        ),
    ]
    assert Ingestion.produces == [
        Artifact(
            name="ingestion_report",
            base="outputs",
            relative="030_ingestion/ingestion_report.json",
        ),
    ]

    # No sobreescribe run(): usa el template method heredado de Flow.
    assert "run" not in vars(Ingestion)
    assert Ingestion.run is Flow.run

    # Completa las 4 fases del template method (definidas explicitamente en
    # la propia clase, no solo heredadas en silencio).
    for hook in ("load_inputs", "validate", "execute", "write_outputs"):
        assert hook in vars(Ingestion), (
            f"Ingestion debe sobreescribir explicitamente el hook {hook!r} "
            "(spec Interfaces / Firmas Publicas, CA-20)."
        )


_INVENTARIO_HEADER = "fecha;sede;clase;stock"
_INVENTARIO_ROWS = [
    "2024-01-01;Sede Centro;Agua 600ml;120",
    "2024-01-02;Sede Norte;Cola 1.5L;80",
]


def _contract_data_inventario_2024() -> dict:
    """DS-ING-8: fuente de los archivos esperados. Un unico dataset
    "inventario" con un unico archivo "inventario_2024.txt" (caso 4,
    subconjunto DS-ING-7 aislado para el separador ';')."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "historical_data": {
            "datasets": [
                {
                    "kind": "inventario",
                    "source_medium": "csv",
                    "periodicity": "mensual",
                    "files": [
                        {
                            "name": "inventario_2024.txt",
                            "period_start": "2023-01-01",
                            "period_end": "2025-12-31",
                        }
                    ],
                }
            ]
        },
    }


def _map_client_data_inventario_2024() -> dict:
    """DS-ING-8: fuente de las columnas esperadas del dataset "inventario"
    (fields[] con name/required), emparejadas por kind con el dataset
    homologo del contrato. Coherente con _contract_data_inventario_2024
    (mismo kind "inventario")."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "datasets": [
            {
                "kind": "inventario",
                "fields": [
                    {"name": "fecha", "required": True},
                    {"name": "sede", "required": True},
                    {"name": "clase", "required": True},
                    {"name": "stock", "required": True},
                ],
            }
        ],
    }


def _build_ctx_fixture_inventario_2024(tmp_path: Path) -> ClientContext:
    """DS-ING-7 (subconjunto aislado, caso 4): ClientContext con
    contract_data.json + map_client_data.json coherentes entre si (dataset
    "inventario" unico) y el archivo crudo inventario_2024.txt (separador
    ';')."""
    return _build_ctx(
        tmp_path,
        _contract_data_inventario_2024(),
        _map_client_data_inventario_2024(),
        {"inventario_2024.txt": "\n".join([_INVENTARIO_HEADER, *_INVENTARIO_ROWS]) + "\n"},
    )


def test_reporte_registra_rows_columns_y_separator_correctos_para_inventario_2024_txt_punto_y_coma(
    tmp_path: Path,
) -> None:
    """Caso 4 (CA-02): para el archivo delimitado por punto y coma
    (inventario_2024.txt) el reporte registra el numero correcto de rows
    (filas de datos, sin cabecera: len(_INVENTARIO_ROWS) == 2) y columns
    (columnas de la cabecera: len(_INVENTARIO_HEADER.split(";")) == 4), y
    separator == ";". La extension no determina el separador (DS-ING-7):
    el archivo es .txt pero el contenido esta delimitado por ';'."""
    ctx = _build_ctx_fixture_inventario_2024(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivo = reporte["datasets"][0]["files"][0]
    assert archivo["name"] == "inventario_2024.txt"
    assert archivo["rows"] == len(_INVENTARIO_ROWS)
    assert archivo["columns"] == len(_INVENTARIO_HEADER.split(";"))
    assert archivo["separator"] == ";"

    assert result.success is True


_INVENTARIO_2025_HEADER = "fecha|sede|clase|stock"
_INVENTARIO_2025_ROWS = [
    "2025-01-01|Sede Centro|Agua 600ml|150",
    "2025-01-02|Sede Norte|Cola 1.5L|95",
]


def _contract_data_inventario_2025() -> dict:
    """DS-ING-8: fuente de los archivos esperados. Un unico dataset
    "inventario" con un unico archivo "inventario_2025.csv" (caso 5,
    subconjunto DS-ING-7 aislado para el separador '|')."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "historical_data": {
            "datasets": [
                {
                    "kind": "inventario",
                    "source_medium": "csv",
                    "periodicity": "mensual",
                    "files": [
                        {
                            "name": "inventario_2025.csv",
                            "period_start": "2023-01-01",
                            "period_end": "2025-12-31",
                        }
                    ],
                }
            ]
        },
    }


def _map_client_data_inventario_2025() -> dict:
    """DS-ING-8: fuente de las columnas esperadas del dataset "inventario"
    (fields[] con name/required), emparejadas por kind con el dataset
    homologo del contrato. Coherente con _contract_data_inventario_2025
    (mismo kind "inventario")."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "datasets": [
            {
                "kind": "inventario",
                "fields": [
                    {"name": "fecha", "required": True},
                    {"name": "sede", "required": True},
                    {"name": "clase", "required": True},
                    {"name": "stock", "required": True},
                ],
            }
        ],
    }


def _build_ctx_fixture_inventario_2025(tmp_path: Path) -> ClientContext:
    """DS-ING-7 (subconjunto aislado, caso 5): ClientContext con
    contract_data.json + map_client_data.json coherentes entre si (dataset
    "inventario" unico) y el archivo crudo inventario_2025.csv (separador
    '|')."""
    return _build_ctx(
        tmp_path,
        _contract_data_inventario_2025(),
        _map_client_data_inventario_2025(),
        {
            "inventario_2025.csv": "\n".join(
                [_INVENTARIO_2025_HEADER, *_INVENTARIO_2025_ROWS]
            )
            + "\n"
        },
    )


def test_reporte_registra_rows_columns_y_separator_correctos_para_inventario_2025_csv_barra_vertical(
    tmp_path: Path,
) -> None:
    """Caso 5 (CA-03): para el archivo delimitado por barra vertical
    (inventario_2025.csv) el reporte registra el numero correcto de rows
    (filas de datos, sin cabecera: len(_INVENTARIO_2025_ROWS) == 2) y columns
    (columnas de la cabecera: len(_INVENTARIO_2025_HEADER.split("|")) == 4),
    y separator == "|"."""
    ctx = _build_ctx_fixture_inventario_2025(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivo = reporte["datasets"][0]["files"][0]
    assert archivo["name"] == "inventario_2025.csv"
    assert archivo["rows"] == len(_INVENTARIO_2025_ROWS)
    assert archivo["columns"] == len(_INVENTARIO_2025_HEADER.split("|"))
    assert archivo["separator"] == "|"

    assert result.success is True


def test_reporte_registra_rows_columns_y_separator_correctos_para_ventas_csv_coma(
    tmp_path: Path,
) -> None:
    """Caso 3 (CA-01): para el archivo delimitado por coma (ventas.csv) el
    reporte registra el numero correcto de rows (filas de datos, sin
    cabecera: len(_VENTAS_ROWS) == 3) y columns (columnas de la cabecera:
    len(_VENTAS_HEADER.split(",")) == 5), y separator == ",". Reutiliza el
    fixture minimo (DS-ING-7, casos 1-3, un unico dataset "ventas")."""
    ctx = _build_ctx_fixture_minimo(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivo = reporte["datasets"][0]["files"][0]
    assert archivo["name"] == "ventas.csv"
    assert archivo["rows"] == len(_VENTAS_ROWS)
    assert archivo["columns"] == len(_VENTAS_HEADER.split(","))
    assert archivo["separator"] == ","

    assert result.success is True


_PRECIOS_HEADER = ["clase", "precio", "moneda"]
_PRECIOS_ROWS = [
    ["Agua 600ml", 1200, "COP"],
    ["Cola 1.5L", 2500, "COP"],
    ["Papas 45g", 900, "COP"],
]


def _contract_data_precios() -> dict:
    """DS-ING-8: fuente de los archivos esperados. Un unico dataset
    "precios" con un unico archivo "precios.xlsx" (caso 6, subconjunto
    DS-ING-7 aislado para el lector .xlsx, source_medium "xlsx")."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "historical_data": {
            "datasets": [
                {
                    "kind": "precios",
                    "source_medium": "xlsx",
                    "periodicity": "mensual",
                    "files": [
                        {
                            "name": "precios.xlsx",
                            "period_start": "2023-01-01",
                            "period_end": "2025-12-31",
                        }
                    ],
                }
            ]
        },
    }


def _map_client_data_precios() -> dict:
    """DS-ING-8: fuente de las columnas esperadas del dataset "precios"
    (fields[] con name/required), emparejadas por kind con el dataset
    homologo del contrato. Coherente con _contract_data_precios (mismo kind
    "precios")."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "datasets": [
            {
                "kind": "precios",
                "fields": [
                    {"name": "clase", "required": True},
                    {"name": "precio", "required": True},
                    {"name": "moneda", "required": True},
                ],
            }
        ],
    }


def _precios_xlsx_bytes() -> bytes:
    """Fabrica en memoria un .xlsx (openpyxl) con cabecera _PRECIOS_HEADER
    y filas _PRECIOS_ROWS en la primera hoja (unica hoja del workbook), y
    devuelve su contenido binario para escribirlo bajo el landing."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(_PRECIOS_HEADER)
    for row in _PRECIOS_ROWS:
        sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_ctx_fixture_precios(tmp_path: Path) -> ClientContext:
    """DS-ING-7 (subconjunto aislado, caso 6): ClientContext con
    contract_data.json + map_client_data.json coherentes entre si (dataset
    "precios" unico) y el archivo crudo precios.xlsx (primera hoja,
    formato Excel, sin separador delimitado)."""
    return _build_ctx(
        tmp_path,
        _contract_data_precios(),
        _map_client_data_precios(),
        {"precios.xlsx": _precios_xlsx_bytes()},
    )


def test_reporte_registra_rows_columns_correctos_y_separator_null_para_precios_xlsx(
    tmp_path: Path,
) -> None:
    """Caso 6 (CA-04): para el archivo Excel (precios.xlsx, primera hoja)
    el reporte registra el numero correcto de rows (filas de datos, sin
    cabecera: len(_PRECIOS_ROWS) == 3) y columns (columnas de la cabecera:
    len(_PRECIOS_HEADER) == 3), y separator == null (None en Python)."""
    ctx = _build_ctx_fixture_precios(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivo = reporte["datasets"][0]["files"][0]
    assert archivo["name"] == "precios.xlsx"
    assert archivo["rows"] == len(_PRECIOS_ROWS)
    assert archivo["columns"] == len(_PRECIOS_HEADER)
    assert archivo["separator"] is None

    assert result.success is True


def _contract_data_completo() -> dict:
    """DS-ING-7/DS-ING-8: fixture completo (fuente de los archivos
    esperados), 3 datasets ("ventas", "inventario" con 2 archivos,
    "precios"), 4 archivos en total. Reutiliza los kinds/archivos ya
    definidos en los fixtures aislados de los casos 3-6 (mismos nombres y
    separadores)."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "historical_data": {
            "datasets": [
                {
                    "kind": "ventas",
                    "source_medium": "csv",
                    "periodicity": "mensual",
                    "files": [
                        {
                            "name": "ventas.csv",
                            "period_start": "2023-01-01",
                            "period_end": "2025-12-31",
                        }
                    ],
                },
                {
                    "kind": "inventario",
                    "source_medium": "csv",
                    "periodicity": "mensual",
                    "files": [
                        {
                            "name": "inventario_2024.txt",
                            "period_start": "2023-01-01",
                            "period_end": "2025-12-31",
                        },
                        {
                            "name": "inventario_2025.csv",
                            "period_start": "2023-01-01",
                            "period_end": "2025-12-31",
                        },
                    ],
                },
                {
                    "kind": "precios",
                    "source_medium": "xlsx",
                    "periodicity": "mensual",
                    "files": [
                        {
                            "name": "precios.xlsx",
                            "period_start": "2023-01-01",
                            "period_end": "2025-12-31",
                        }
                    ],
                },
            ]
        },
    }


def _map_client_data_completo() -> dict:
    """DS-ING-8: fuente de las columnas esperadas de los 3 datasets del
    fixture completo, coherente con _contract_data_completo (mismos
    kinds). Reutiliza los fields[] ya definidos en los fixtures aislados de
    los casos 3-6."""
    return {
        "schema_version": "0.1",
        "client": {"code": "ABC", "name": "Cliente ABC S.A.", "sector": "retail"},
        "datasets": [
            _map_client_data_minimo()["datasets"][0],
            _map_client_data_inventario_2024()["datasets"][0],
            _map_client_data_precios()["datasets"][0],
        ],
    }


def _build_ctx_fixture_completo(
    tmp_path: Path, precios_xlsx: bytes | None = None
) -> ClientContext:
    """DS-ING-7 (fixture completo, caso 7): ClientContext con
    contract_data.json + map_client_data.json coherentes entre si (3
    datasets) y los 4 archivos crudos del landing (ventas.csv coma,
    inventario_2024.txt ';', inventario_2025.csv '|', precios.xlsx).

    precios_xlsx permite inyectar los mismos bytes de xlsx en dos ctx
    equivalentes; openpyxl embebe timestamps al serializar, de modo que
    dos llamadas independientes a _precios_xlsx_bytes() producen bytes
    distintos. El test de determinismo (CA-13) requiere origen identico
    para aislar el determinismo de la copia en bronze. Por defecto genera
    un xlsx fresco (comportamiento previo, suficiente para el resto)."""
    return _build_ctx(
        tmp_path,
        _contract_data_completo(),
        _map_client_data_completo(),
        {
            "ventas.csv": "\n".join([_VENTAS_HEADER, *_VENTAS_ROWS]) + "\n",
            "inventario_2024.txt": "\n".join(
                [_INVENTARIO_HEADER, *_INVENTARIO_ROWS]
            )
            + "\n",
            "inventario_2025.csv": "\n".join(
                [_INVENTARIO_2025_HEADER, *_INVENTARIO_2025_ROWS]
            )
            + "\n",
            "precios.xlsx": precios_xlsx
            if precios_xlsx is not None
            else _precios_xlsx_bytes(),
        },
    )


def test_reporte_expone_por_archivo_name_rows_y_columns_para_los_4_archivos_del_fixture_completo(
    tmp_path: Path,
) -> None:
    """Caso 7 (CA-15): sobre el fixture completo (DS-ING-7, 3 datasets, 4
    archivos con distintos separadores/formato) el reporte expone, por
    CADA archivo, los campos name/rows/columns con los valores correctos.
    A diferencia de los casos 3-6 (que aislaban un unico dataset/archivo
    por fixture), este caso ejercita varios datasets con mas de un archivo
    por dataset ("inventario" con 2 archivos) para confirmar que ningun
    archivo del reporte queda sin su name/rows/columns."""
    ctx = _build_ctx_fixture_completo(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    esperado_por_nombre = {
        "ventas.csv": (len(_VENTAS_ROWS), len(_VENTAS_HEADER.split(","))),
        "inventario_2024.txt": (
            len(_INVENTARIO_ROWS),
            len(_INVENTARIO_HEADER.split(";")),
        ),
        "inventario_2025.csv": (
            len(_INVENTARIO_2025_ROWS),
            len(_INVENTARIO_2025_HEADER.split("|")),
        ),
        "precios.xlsx": (len(_PRECIOS_ROWS), len(_PRECIOS_HEADER)),
    }

    archivos_vistos = []
    for dataset in reporte["datasets"]:
        for archivo in dataset["files"]:
            archivos_vistos.append(archivo["name"])
            assert "name" in archivo
            assert "rows" in archivo
            assert "columns" in archivo
            rows_esperadas, columns_esperadas = esperado_por_nombre[archivo["name"]]
            assert archivo["rows"] == rows_esperadas
            assert archivo["columns"] == columns_esperadas

    assert sorted(archivos_vistos) == sorted(esperado_por_nombre.keys())
    assert result.success is True


def test_cada_archivo_valido_tiene_copia_byte_a_byte_identica_en_bronze(
    tmp_path: Path,
) -> None:
    """Caso 8 (CA-11): sobre el fixture completo (DS-ING-7, 3 datasets, 4
    archivos validos: ventas.csv coma, inventario_2024.txt ';',
    inventario_2025.csv '|', precios.xlsx), para CADA archivo existe en
    ctx.bronze_dir/<name> una copia byte a byte identica al original del
    landing (ctx.inputs_dir/"030_ingestion"/<name>). Se compara con
    Path.read_bytes() para verificar fidelidad binaria exacta, no solo
    contenido de texto."""
    ctx = _build_ctx_fixture_completo(tmp_path)
    landing_dir = ctx.inputs_dir / "030_ingestion"

    flow = Ingestion()
    flow.run(ctx)

    for name in (
        "ventas.csv",
        "inventario_2024.txt",
        "inventario_2025.csv",
        "precios.xlsx",
    ):
        origen = landing_dir / name
        destino = ctx.bronze_dir / name
        assert destino.exists(), f"falta la copia en bronze de {name!r}"
        assert destino.read_bytes() == origen.read_bytes()


def test_copia_en_bronze_conserva_formato_separador_y_extension(
    tmp_path: Path,
) -> None:
    """Caso 9 (CA-12): la copia en ctx.bronze_dir conserva formato,
    separador y extension del original, no solo bytes identicos en
    abstracto (ya cubierto por el caso 8). Dos focos concretos:
    - inventario_2025.csv (separador '|'): la copia en bronze sigue
      delimitada por '|' (su cabecera, leida desde bronze y partida por
      '|', produce las mismas 4 columnas que el original); la extension
      .csv se conserva.
    - precios.xlsx: la copia en bronze conserva la extension .xlsx y es
      un .xlsx valido y abrible con openpyxl, con la misma cabecera y
      filas que el original (no se re-serializa a otro formato)."""
    ctx = _build_ctx_fixture_completo(tmp_path)

    flow = Ingestion()
    flow.run(ctx)

    # Delimitado por '|': la copia en bronze sigue separada por '|'.
    copia_barra = ctx.bronze_dir / "inventario_2025.csv"
    assert copia_barra.suffix == ".csv"
    lineas_copia = [
        line
        for line in copia_barra.read_text(encoding="utf-8").splitlines()
        if line.strip() != ""
    ]
    cabecera_copia = lineas_copia[0].split("|")
    assert cabecera_copia == _INVENTARIO_2025_HEADER.split("|")
    assert len(cabecera_copia) == 4
    # No debe haberse re-separado por coma u otro delimitador.
    assert "," not in lineas_copia[0]

    # .xlsx: la copia conserva extension y es un .xlsx valido, byte-identico.
    copia_xlsx = ctx.bronze_dir / "precios.xlsx"
    assert copia_xlsx.suffix == ".xlsx"
    origen_xlsx = ctx.inputs_dir / "030_ingestion" / "precios.xlsx"
    assert copia_xlsx.read_bytes() == origen_xlsx.read_bytes()

    workbook = openpyxl.load_workbook(copia_xlsx, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    filas = [
        row
        for row in sheet.iter_rows(values_only=True)
        if any(cell is not None for cell in row)
    ]
    assert list(filas[0]) == _PRECIOS_HEADER
    assert [list(row) for row in filas[1:]] == _PRECIOS_ROWS


def test_presentes_igual_declarados_sin_missing_ni_unexpected_files_ingested_igual_declared(
    tmp_path: Path,
) -> None:
    """Caso 10 (CA-05): sobre el fixture completo (DS-ING-7), donde los
    archivos presentes en el landing coinciden EXACTAMENTE con los
    declarados en contract_data.json (historical_data.datasets[].files[].
    name, DS-ING-8), el reporte no registra ninguna inconsistencia
    missing_file ni unexpected_file (en ningun archivo de ningun dataset),
    unexpected_files == [], y summary.files_ingested ==
    summary.files_declared."""
    ctx = _build_ctx_fixture_completo(tmp_path)

    flow = Ingestion()
    flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    tipos_inconsistencias = [
        inconsistencia["type"]
        for dataset in reporte["datasets"]
        for archivo in dataset["files"]
        for inconsistencia in archivo["inconsistencies"]
    ]
    assert "missing_file" not in tipos_inconsistencias
    assert "unexpected_file" not in tipos_inconsistencias
    assert reporte["unexpected_files"] == []
    assert (
        reporte["summary"]["files_ingested"] == reporte["summary"]["files_declared"]
    )


def _build_ctx_fixture_archivo_faltante(tmp_path: Path) -> ClientContext:
    """DS-ING-7/DS-ING-8 (variante "Faltante", caso 11): mismo contrato +
    mapa completos que _build_ctx_fixture_completo (3 datasets, 4 archivos
    declarados en contract_data.json), pero el landing NO deposita
    "inventario_2025.csv" (declarado en el dataset "inventario" del
    contrato). Los otros 3 archivos si se depositan, para aislar el efecto
    del faltante."""
    return _build_ctx(
        tmp_path,
        _contract_data_completo(),
        _map_client_data_completo(),
        {
            "ventas.csv": "\n".join([_VENTAS_HEADER, *_VENTAS_ROWS]) + "\n",
            "inventario_2024.txt": "\n".join(
                [_INVENTARIO_HEADER, *_INVENTARIO_ROWS]
            )
            + "\n",
            "precios.xlsx": _precios_xlsx_bytes(),
        },
    )


def test_archivo_declarado_en_contrato_no_presente_marca_missing_sin_copiar_y_success_false(
    tmp_path: Path,
) -> None:
    """Caso 11 (CA-06): "inventario_2025.csv" esta declarado en
    contract_data.json (dataset "inventario", DS-ING-8) pero no se deposita
    en el landing. El reporte debe marcar ese archivo con status=="missing"
    y registrar una inconsistencia de type "missing_file"; no debe existir
    copia en ctx.bronze_dir/"inventario_2025.csv"; y FlowResult.success debe
    ser False (hay al menos una inconsistencia)."""
    ctx = _build_ctx_fixture_archivo_faltante(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivo_faltante = None
    for dataset in reporte["datasets"]:
        for archivo in dataset["files"]:
            if archivo["name"] == "inventario_2025.csv":
                archivo_faltante = archivo
    assert archivo_faltante is not None, (
        "inventario_2025.csv debe seguir apareciendo en el reporte "
        "(declarado en el contrato) aunque no este presente en el landing"
    )
    assert archivo_faltante["status"] == "missing"
    tipos_inconsistencias = [
        inconsistencia["type"]
        for inconsistencia in archivo_faltante["inconsistencies"]
    ]
    assert "missing_file" in tipos_inconsistencias

    assert not (ctx.bronze_dir / "inventario_2025.csv").exists()

    assert result.success is False


def _build_ctx_fixture_archivos_sobrantes(tmp_path: Path) -> ClientContext:
    """DS-ING-7/DS-ING-8 (variante "Sobrante", caso 12): mismo contrato +
    mapa completos que _build_ctx_fixture_completo (3 datasets, 4 archivos
    declarados en contract_data.json), mas dos archivos adicionales en el
    landing que NO figuran en ningun historical_data.datasets[].files[].name
    del contrato ("zzz_sobrante.csv" y "aaa_sobrante.txt", en orden no
    alfabetico al depositarlos para ejercitar el orden alfabetico ascendente
    exigido por DS-ING-6 al reportarlos en unexpected_files)."""
    return _build_ctx(
        tmp_path,
        _contract_data_completo(),
        _map_client_data_completo(),
        {
            "ventas.csv": "\n".join([_VENTAS_HEADER, *_VENTAS_ROWS]) + "\n",
            "inventario_2024.txt": "\n".join(
                [_INVENTARIO_HEADER, *_INVENTARIO_ROWS]
            )
            + "\n",
            "inventario_2025.csv": "\n".join(
                [_INVENTARIO_2025_HEADER, *_INVENTARIO_2025_ROWS]
            )
            + "\n",
            "precios.xlsx": _precios_xlsx_bytes(),
            "zzz_sobrante.csv": "col_a,col_b\n1,2\n",
            "aaa_sobrante.txt": "algo\n",
        },
    )


def test_archivo_presente_no_declarado_en_contrato_va_a_unexpected_files_sin_copiar_y_success_false(
    tmp_path: Path,
) -> None:
    """Caso 12 (CA-07): "zzz_sobrante.csv" y "aaa_sobrante.txt" estan
    presentes en el landing pero no figuran declarados en ningun
    historical_data.datasets[].files[].name de contract_data.json
    (DS-ING-8). Ambos deben aparecer en reporte["unexpected_files"] en
    orden alfabetico ascendente (DS-ING-6); ninguno debe copiarse a
    ctx.bronze_dir; y FlowResult.success debe ser False (hay al menos una
    inconsistencia unexpected_file, CA-07)."""
    ctx = _build_ctx_fixture_archivos_sobrantes(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    assert reporte["unexpected_files"] == ["aaa_sobrante.txt", "zzz_sobrante.csv"]

    assert not (ctx.bronze_dir / "zzz_sobrante.csv").exists()
    assert not (ctx.bronze_dir / "aaa_sobrante.txt").exists()

    assert result.success is False


def _build_ctx_fixture_columna_requerida_ausente(tmp_path: Path) -> ClientContext:
    """DS-ING-7/DS-ING-8 (variante "columna requerida ausente", caso 13):
    mismo contrato + mapa completos que _build_ctx_fixture_completo (3
    datasets, 4 archivos), pero "ventas.csv" se deposita en el landing SIN
    la columna "clase" (required=True segun el dataset homologo "ventas"
    de map_client_data.json, DS-ING-8). Los otros 3 archivos se depositan
    sin cambios, para aislar el efecto de la columna faltante en
    ventas.csv."""
    ventas_header_sin_clase = "fecha,sede,cantidad,precio_unitario"
    ventas_rows_sin_clase = [
        "2024-01-01,Sede Centro,10,1200",
        "2024-01-02,Sede Norte,5,2500",
        "2024-01-03,Sede Centro,20,900",
    ]
    return _build_ctx(
        tmp_path,
        _contract_data_completo(),
        _map_client_data_completo(),
        {
            "ventas.csv": "\n".join(
                [ventas_header_sin_clase, *ventas_rows_sin_clase]
            )
            + "\n",
            "inventario_2024.txt": "\n".join(
                [_INVENTARIO_HEADER, *_INVENTARIO_ROWS]
            )
            + "\n",
            "inventario_2025.csv": "\n".join(
                [_INVENTARIO_2025_HEADER, *_INVENTARIO_2025_ROWS]
            )
            + "\n",
            "precios.xlsx": _precios_xlsx_bytes(),
        },
    )


def test_archivo_sin_columna_requerida_segun_mapa_marca_rejected_missing_column_sin_copiar_y_success_false(
    tmp_path: Path,
) -> None:
    """Caso 13 (CA-08): "ventas.csv" esta presente en el landing pero le
    falta la columna "clase", declarada con required=True en el dataset
    homologo "ventas" de map_client_data.json (emparejado por kind,
    DS-ING-8). El reporte debe marcar ese archivo con status=="rejected" y
    registrar una inconsistencia de type "missing_column"; no debe existir
    copia en ctx.bronze_dir/"ventas.csv"; y FlowResult.success debe ser
    False (hay al menos una inconsistencia)."""
    ctx = _build_ctx_fixture_columna_requerida_ausente(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivo_ventas = None
    for dataset in reporte["datasets"]:
        for archivo in dataset["files"]:
            if archivo["name"] == "ventas.csv":
                archivo_ventas = archivo
    assert archivo_ventas is not None, (
        "ventas.csv debe seguir apareciendo en el reporte (presente en el "
        "landing) aunque le falte una columna requerida segun el mapa"
    )
    assert archivo_ventas["status"] == "rejected"
    tipos_inconsistencias = [
        inconsistencia["type"] for inconsistencia in archivo_ventas["inconsistencies"]
    ]
    assert "missing_column" in tipos_inconsistencias

    assert not (ctx.bronze_dir / "ventas.csv").exists()

    assert result.success is False


def _build_ctx_fixture_columna_no_declarada(tmp_path: Path) -> ClientContext:
    """DS-ING-7/DS-ING-8 (variante "columna no declarada", caso 14): mismo
    contrato + mapa completos que _build_ctx_fixture_completo (3 datasets,
    4 archivos), pero "ventas.csv" se deposita en el landing con la
    columna "clase" renombrada a "categoria" (fuera de los fields[] del
    dataset homologo "ventas" de map_client_data.json, DS-ING-8). Los
    otros 3 archivos se depositan sin cambios, para aislar el efecto de la
    columna renombrada en ventas.csv (plan.md, variante "Columna no
    declarada", caso 14)."""
    ventas_header_categoria = "fecha,sede,categoria,cantidad,precio_unitario"
    return _build_ctx(
        tmp_path,
        _contract_data_completo(),
        _map_client_data_completo(),
        {
            "ventas.csv": "\n".join([ventas_header_categoria, *_VENTAS_ROWS]) + "\n",
            "inventario_2024.txt": "\n".join(
                [_INVENTARIO_HEADER, *_INVENTARIO_ROWS]
            )
            + "\n",
            "inventario_2025.csv": "\n".join(
                [_INVENTARIO_2025_HEADER, *_INVENTARIO_2025_ROWS]
            )
            + "\n",
            "precios.xlsx": _precios_xlsx_bytes(),
        },
    )


def test_archivo_con_columna_no_declarada_segun_mapa_marca_rejected_unexpected_column_sin_copiar_y_success_false(
    tmp_path: Path,
) -> None:
    """Caso 14 (CA-09): "ventas.csv" esta presente en el landing pero su
    columna "clase" fue renombrada a "categoria", que no figura en los
    fields[] del dataset homologo "ventas" de map_client_data.json
    (emparejado por kind, DS-ING-8). El reporte debe marcar ese archivo
    con status=="rejected" y registrar una inconsistencia de type
    "unexpected_column"; no debe existir copia en
    ctx.bronze_dir/"ventas.csv"; y FlowResult.success debe ser False (hay
    al menos una inconsistencia)."""
    ctx = _build_ctx_fixture_columna_no_declarada(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivo_ventas = None
    for dataset in reporte["datasets"]:
        for archivo in dataset["files"]:
            if archivo["name"] == "ventas.csv":
                archivo_ventas = archivo
    assert archivo_ventas is not None, (
        "ventas.csv debe seguir apareciendo en el reporte (presente en el "
        "landing) aunque tenga una columna no declarada en el mapa"
    )
    assert archivo_ventas["status"] == "rejected"
    tipos_inconsistencias = [
        inconsistencia["type"] for inconsistencia in archivo_ventas["inconsistencies"]
    ]
    assert "unexpected_column" in tipos_inconsistencias

    assert not (ctx.bronze_dir / "ventas.csv").exists()

    assert result.success is False


def _build_ctx_fixture_columna_opcional_ausente(tmp_path: Path) -> ClientContext:
    """DS-ING-7/DS-ING-8 (variante "columna opcional ausente", caso 15):
    mismo contrato + mapa minimos que _build_ctx_fixture_minimo (un unico
    dataset "ventas" con un unico archivo "ventas.csv"), pero "ventas.csv"
    se deposita en el landing SIN la columna "precio_unitario", declarada
    con required=False en el dataset homologo "ventas" de
    map_client_data.json (emparejado por kind, DS-ING-8; plan.md, variante
    "Columna opcional ausente", caso 15)."""
    ventas_header_sin_precio_unitario = "fecha,sede,clase,cantidad"
    ventas_rows_sin_precio_unitario = [
        "2024-01-01,Sede Centro,Agua 600ml,10",
        "2024-01-02,Sede Norte,Cola 1.5L,5",
        "2024-01-03,Sede Centro,Papas 45g,20",
    ]
    return _build_ctx(
        tmp_path,
        _contract_data_minimo(),
        _map_client_data_minimo(),
        {
            "ventas.csv": "\n".join(
                [ventas_header_sin_precio_unitario, *ventas_rows_sin_precio_unitario]
            )
            + "\n",
        },
    )


def test_archivo_sin_columna_opcional_segun_mapa_no_es_inconsistencia_y_queda_ingested(
    tmp_path: Path,
) -> None:
    """Caso 15 (CA-10): "ventas.csv" esta presente en el landing pero le
    falta la columna "precio_unitario", declarada con required=False en el
    dataset homologo "ventas" de map_client_data.json (emparejado por kind,
    DS-ING-8). Al ser una columna opcional (no requerida) su ausencia NO es
    una inconsistencia (_validate_columns, regla (c), plan.md linea 113):
    el archivo debe quedar con status=="ingested", sin ninguna
    inconsistencia de type "missing_column"/"unexpected_column", debe
    existir su copia byte a byte en ctx.bronze_dir/"ventas.csv", y
    FlowResult.success debe ser True (unico archivo del fixture, sin
    ninguna otra novedad)."""
    ctx = _build_ctx_fixture_columna_opcional_ausente(tmp_path)

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivo_ventas = None
    for dataset in reporte["datasets"]:
        for archivo in dataset["files"]:
            if archivo["name"] == "ventas.csv":
                archivo_ventas = archivo
    assert archivo_ventas is not None, (
        "ventas.csv debe seguir apareciendo en el reporte (presente en el "
        "landing) aunque le falte una columna opcional segun el mapa"
    )
    assert archivo_ventas["status"] == "ingested"
    assert archivo_ventas["inconsistencies"] == []
    assert archivo_ventas["columns"] == 4

    ruta_bronze = ctx.bronze_dir / "ventas.csv"
    assert ruta_bronze.exists()
    assert ruta_bronze.read_bytes() == (
        ctx.inputs_dir / "030_ingestion/ventas.csv"
    ).read_bytes()

    assert result.success is True


def _build_ctx_fixture_multiples_inconsistencias(tmp_path: Path) -> ClientContext:
    """DS-ING-7/DS-ING-9 (variante "multiples inconsistencias", caso 16):
    mismo contrato + mapa completos que _build_ctx_fixture_completo (3
    datasets, 4 archivos declarados), combinando cuatro mutaciones
    simultaneas para provocar los 4 tipos del vocabulario cerrado
    (DS-ING-9) en un unico run:
    - "inventario_2025.csv" (declarado en el contrato) NO se deposita en
      el landing -> missing_file.
    - "zzz_sobrante.csv" se deposita sin estar declarado en el contrato
      -> unexpected_file.
    - "ventas.csv" se deposita sin la columna "clase" (required=True
      segun el dataset homologo "ventas" del mapa, ver caso 13) ->
      missing_column.
    - "inventario_2024.txt" se deposita con la columna "clase" renombrada
      a "categoria" (fuera de los fields[] del dataset homologo
      "inventario" del mapa, ver caso 14) -> unexpected_column (ademas
      dispara missing_column por la ausencia de "clase" en ese mismo
      archivo, lo cual es admisible: el caso solo exige que EXISTAN las
      4 variantes en algun punto del run, no que cada archivo aporte una
      unica variante)."""
    ventas_header_sin_clase = "fecha,sede,cantidad,precio_unitario"
    ventas_rows_sin_clase = [
        "2024-01-01,Sede Centro,10,1200",
        "2024-01-02,Sede Norte,5,2500",
        "2024-01-03,Sede Centro,20,900",
    ]
    inventario_2024_header_categoria = "fecha;sede;categoria;stock"
    return _build_ctx(
        tmp_path,
        _contract_data_completo(),
        _map_client_data_completo(),
        {
            "ventas.csv": "\n".join(
                [ventas_header_sin_clase, *ventas_rows_sin_clase]
            )
            + "\n",
            "inventario_2024.txt": "\n".join(
                [inventario_2024_header_categoria, *_INVENTARIO_ROWS]
            )
            + "\n",
            "precios.xlsx": _precios_xlsx_bytes(),
            "zzz_sobrante.csv": "col_a,col_b\n1,2\n",
        },
    )


def test_lista_top_level_inconsistencies_agrega_tipos_del_vocabulario_cerrado_con_detail_no_vacio(
    tmp_path: Path,
) -> None:
    """Caso 16 (CA-16, DS-ING-9): sobre un escenario con inconsistencias
    de los 4 tipos distintos del vocabulario cerrado (missing_file,
    unexpected_file, missing_column, unexpected_column;
    _build_ctx_fixture_multiples_inconsistencias), la lista top-level
    reporte["inconsistencies"] (enmienda DS-ING-9, ADR D-078):
    - no esta vacia;
    - cada entrada tiene un "type" dentro del vocabulario cerrado de 4;
    - cada entrada tiene un "detail" string no vacio;
    - incluye efectivamente una entrada de type "unexpected_file" (el
      hogar estructural que saldaba la deuda documentada en el caso 12)."""
    ctx = _build_ctx_fixture_multiples_inconsistencias(tmp_path)

    flow = Ingestion()
    flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    inconsistencias = reporte["inconsistencies"]
    assert inconsistencias != []

    vocabulario_cerrado = {
        "missing_file",
        "unexpected_file",
        "missing_column",
        "unexpected_column",
    }
    for inconsistencia in inconsistencias:
        assert inconsistencia["type"] in vocabulario_cerrado
        assert isinstance(inconsistencia["detail"], str)
        assert inconsistencia["detail"] != ""

    tipos = [inconsistencia["type"] for inconsistencia in inconsistencias]
    assert "unexpected_file" in tipos


def test_summary_reporta_los_4_conteos_derivados_del_contrato_y_coherentes_con_el_detalle(
    tmp_path: Path,
) -> None:
    """Caso 17 (CA-17, TSK-31/TSK-10): sobre un escenario con al menos una
    inconsistencia (_build_ctx_fixture_multiples_inconsistencias, reutilizado
    del caso 16: 3 datasets/4 archivos declarados en contract_data.json,
    DS-ING-8, con "inventario_2025.csv" missing, "ventas.csv" y
    "inventario_2024.txt" rejected -missing_column/unexpected_column- y
    "precios.xlsx" ingested; ademas "zzz_sobrante.csv" es unexpected_file,
    NO declarado en el contrato), report["summary"] expone los 4 conteos
    EXACTOS y coherentes con el propio detalle datasets[].files[]:
    - datasets_declared == len(historical_data.datasets) del contrato
      (contract_data.json, DS-ING-8), 3 en este fixture;
    - files_declared == numero total de historical_data.datasets[].files[]
      del contrato (DS-ING-8), 4 en este fixture (no cuenta
      "zzz_sobrante.csv", que no esta declarado);
    - files_ingested == numero de archivos con status=="ingested" en el
      detalle (1: solo "precios.xlsx" en este fixture);
    - files_with_inconsistencies == numero de archivos con status!=
      "ingested" (missing/rejected) en el detalle (3 en este fixture:
      "inventario_2025.csv" missing, "ventas.csv" e "inventario_2024.txt"
      rejected)."""
    ctx = _build_ctx_fixture_multiples_inconsistencias(tmp_path)
    contract = _contract_data_completo()

    flow = Ingestion()
    flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))
    summary = reporte["summary"]

    datasets_contrato = contract["historical_data"]["datasets"]
    datasets_declared_esperado = len(datasets_contrato)
    files_declared_esperado = sum(
        len(dataset["files"]) for dataset in datasets_contrato
    )
    assert datasets_declared_esperado == 3
    assert files_declared_esperado == 4

    todos_los_archivos = [
        file_
        for dataset in reporte["datasets"]
        for file_ in dataset["files"]
    ]
    files_ingested_esperado = sum(
        1 for file_ in todos_los_archivos if file_["status"] == "ingested"
    )
    files_with_inconsistencies_esperado = sum(
        1 for file_ in todos_los_archivos if file_["status"] != "ingested"
    )
    assert files_ingested_esperado == 1
    assert files_with_inconsistencies_esperado == 3
    assert files_ingested_esperado != files_declared_esperado
    assert files_with_inconsistencies_esperado > 0

    assert summary["datasets_declared"] == datasets_declared_esperado
    assert summary["files_declared"] == files_declared_esperado
    assert summary["files_ingested"] == files_ingested_esperado
    assert summary["files_with_inconsistencies"] == files_with_inconsistencies_esperado


def test_success_es_true_sii_reporte_sin_inconsistencias_y_reporte_se_escribe_en_ambos_casos(
    tmp_path: Path,
) -> None:
    """Caso 18 (CA-19, TSK-32/TSK-10): FlowResult.success == True SI Y SOLO
    SI el reporte no registra NINGUNA inconsistencia; en caso contrario
    False, y el reporte (ingestion_report.json) se escribe IGUALMENTE en
    ctx.outputs_dir/"030_ingestion/ingestion_report.json" (DS-ING-1,
    DS-ING-6: el reporte se escribe siempre que se llega a execute, haya o
    no inconsistencias). Cubre las DOS direcciones del "sii":
    - escenario limpio (_build_ctx_fixture_completo, DS-ING-7: 3 datasets/4
      archivos, todos presentes/declarados/con columnas correctas) ->
      reporte["inconsistencies"] == [] y result.success is True; el reporte
      existe en disco.
    - escenario con inconsistencias (_build_ctx_fixture_multiples_
      inconsistencias, DS-ING-9, reutilizado de los casos 16-17: missing_file
      + unexpected_file + missing_column + unexpected_column simultaneos) ->
      reporte["inconsistencies"] != [] y result.success is False; el reporte
      se escribio IGUALMENTE (existe en disco pese al fallo)."""
    ruta_relativa_reporte = Path("030_ingestion/ingestion_report.json")

    # Direccion 1: sin inconsistencias -> success True, reporte escrito.
    ctx_limpio = _build_ctx_fixture_completo(tmp_path / "limpio")
    resultado_limpio = Ingestion().run(ctx_limpio)
    ruta_reporte_limpio = ctx_limpio.outputs_dir / ruta_relativa_reporte
    assert ruta_reporte_limpio.exists()
    reporte_limpio = json.loads(ruta_reporte_limpio.read_text(encoding="utf-8"))
    assert reporte_limpio["inconsistencies"] == []
    assert reporte_limpio["success"] is True
    assert resultado_limpio.success is True

    # Direccion 2: con >= 1 inconsistencia -> success False, reporte
    # escrito IGUALMENTE (no se omite pese al fallo).
    ctx_con_inconsistencias = _build_ctx_fixture_multiples_inconsistencias(
        tmp_path / "con_inconsistencias"
    )
    resultado_con_inconsistencias = Ingestion().run(ctx_con_inconsistencias)
    ruta_reporte_inconsistente = (
        ctx_con_inconsistencias.outputs_dir / ruta_relativa_reporte
    )
    assert ruta_reporte_inconsistente.exists()
    reporte_inconsistente = json.loads(
        ruta_reporte_inconsistente.read_text(encoding="utf-8")
    )
    assert reporte_inconsistente["inconsistencies"] != []
    assert reporte_inconsistente["success"] is False
    assert resultado_con_inconsistencias.success is False


def test_inconsistencia_parcial_copia_los_validos_y_excluye_el_invalido_con_bronze_path_coherente(
    tmp_path: Path,
) -> None:
    """Caso 19 (CA-18, TSK-33/TSK-08/TSK-09, DS-ING-5): sobre un escenario
    con varios archivos donde unos son validos y uno es invalido, se
    reutiliza _build_ctx_fixture_columna_requerida_ausente (mismo fixture
    del caso 13): "ventas.csv" queda "rejected" (falta la columna
    requerida "clase" segun el mapa) y los otros 3 archivos del fixture
    completo ("inventario_2024.txt", "inventario_2025.csv", "precios.xlsx")
    quedan validos sin cambios. DS-ING-5 exige que la unidad de copia sea
    el ARCHIVO: una inconsistencia en "ventas.csv" no debe impedir que los
    3 archivos validos se copien a ctx.bronze_dir. Se verifica:
    - los 3 validos: status=="ingested", bronze_path no nulo en el
      reporte, Y existen fisicamente en ctx.bronze_dir con contenido byte
      a byte identico al original del landing.
    - el invalido ("ventas.csv"): status=="rejected", bronze_path es
      null en el reporte, Y NO existe en ctx.bronze_dir.
    - FlowResult.success is False (hay al menos una inconsistencia)."""
    ctx = _build_ctx_fixture_columna_requerida_ausente(tmp_path)
    landing_dir = ctx.inputs_dir / "030_ingestion"

    flow = Ingestion()
    result = flow.run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    reporte = json.loads(ruta_reporte.read_text(encoding="utf-8"))

    archivos_por_nombre = {
        archivo["name"]: archivo
        for dataset in reporte["datasets"]
        for archivo in dataset["files"]
    }

    nombres_validos = (
        "inventario_2024.txt",
        "inventario_2025.csv",
        "precios.xlsx",
    )
    for name in nombres_validos:
        archivo = archivos_por_nombre[name]
        assert archivo["status"] == "ingested"
        assert archivo["bronze_path"] is not None, (
            f"{name!r} es valido: su entrada en el reporte debe traer "
            "bronze_path no nulo (DS-ING-2/CA-18)"
        )
        destino = ctx.bronze_dir / name
        origen = landing_dir / name
        assert destino.exists(), f"falta la copia en bronze de {name!r}"
        assert destino.read_bytes() == origen.read_bytes()

    archivo_ventas = archivos_por_nombre["ventas.csv"]
    assert archivo_ventas["status"] == "rejected"
    assert archivo_ventas["bronze_path"] is None
    assert not (ctx.bronze_dir / "ventas.csv").exists()

    assert result.success is False


def test_dos_ejecuciones_con_las_mismas_entradas_producen_reporte_y_copias_bronze_byte_identicos(
    tmp_path: Path,
) -> None:
    """Caso 20 (CA-13, TSK-34/TSK-11, DS-ING-6): dos invocaciones de
    Ingestion().run(ctx) con las MISMAS entradas producen un
    ingestion_report.json byte-identico y copias en bronze byte-identicas
    entre ambas corridas.

    Eleccion de diseno (NC-1/NC-6, documentada en vez de asumida en
    silencio): se usan DOS ClientContext EQUIVALENTES (mismo fixture
    DS-ING-7 completo -3 datasets, 4 archivos, todos "ingested"-,
    construido dos veces bajo dos subcarpetas independientes de tmp_path)
    en vez de reutilizar el mismo ctx para ambas corridas. Reutilizar el
    mismo ctx mezclaria el determinismo (CA-13, foco de este caso) con la
    idempotencia de sobrescritura (escribir encima de un reporte/copias ya
    existentes de una corrida previa), que es un comportamiento distinto no
    pedido por este caso. Con dos ctx equivalentes, cada corrida escribe
    sobre un arbol de archivos limpio, y la comparacion byte a byte aisla
    exclusivamente si el proceso de serializacion/copia es determinista
    (mismo input -> mismo output), tal como exige DS-ING-6 (sort_keys=True
    + indent=2 + newline final + orden estable) para el reporte y la copia
    fiel sin re-serializar para bronze."""
    precios_xlsx = _precios_xlsx_bytes()
    ctx_1 = _build_ctx_fixture_completo(tmp_path / "corrida_1", precios_xlsx)
    ctx_2 = _build_ctx_fixture_completo(tmp_path / "corrida_2", precios_xlsx)

    Ingestion().run(ctx_1)
    Ingestion().run(ctx_2)

    ruta_reporte_1 = ctx_1.outputs_dir / "030_ingestion/ingestion_report.json"
    ruta_reporte_2 = ctx_2.outputs_dir / "030_ingestion/ingestion_report.json"
    assert ruta_reporte_1.read_bytes() == ruta_reporte_2.read_bytes()

    for name in (
        "ventas.csv",
        "inventario_2024.txt",
        "inventario_2025.csv",
        "precios.xlsx",
    ):
        copia_1 = (ctx_1.bronze_dir / name).read_bytes()
        copia_2 = (ctx_2.bronze_dir / name).read_bytes()
        assert copia_1 == copia_2, f"copia en bronze de {name!r} no es byte-identica entre corridas"


def test_contract_data_ausente_lanza_flow_contract_error_en_validate_sin_reporte_ni_copia(
    tmp_path: Path,
) -> None:
    """Caso 21 (CA-21, TSK-35/TSK-03, DS-ING-1): si contract_data.json (uno
    de los requires del flujo) esta AUSENTE del disco, Ingestion().run(ctx)
    lanza FlowContractError en la fase validate (heredada de Flow base, sin
    sobreescritura de contenido: Ingestion.validate() solo delega en
    super().validate(ctx), caso 2/CA-20), ANTES de tocar bronze. Se verifica
    con pytest.raises(FlowContractError) y se confirma que NO hay ninguna
    salida parcial: ni ingestion_report.json (produces[0].path(ctx)) ni
    copia alguna en ctx.bronze_dir (que ya existe vacio por el scaffold de
    create_client, DS-ING-1: contrato de errores, solo requires ausente
    aborta; las inconsistencias de datos nunca abortan)."""
    ctx = _build_ctx_fixture_completo(tmp_path)
    contrato_path = ctx.outputs_dir / "010_discovery/contract_data.json"
    contrato_path.unlink()

    with pytest.raises(FlowContractError):
        Ingestion().run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    assert not ruta_reporte.exists()
    assert list(ctx.bronze_dir.iterdir()) == []


def test_map_client_data_ausente_lanza_flow_contract_error_en_validate_sin_reporte_ni_copia(
    tmp_path: Path,
) -> None:
    """Caso 22 (CA-21, TSK-36/TSK-03, DS-ING-1): analogo al caso 21, pero con
    map_client_data.json (el OTRO requires del flujo) AUSENTE del disco en
    vez de contract_data.json. Ingestion().run(ctx) lanza FlowContractError
    en la fase validate (heredada de Flow base, sin sobreescritura de
    contenido: Ingestion.validate() solo delega en super().validate(ctx),
    caso 2/CA-20), ANTES de tocar bronze. Se verifica con
    pytest.raises(FlowContractError) y se confirma que NO hay ninguna salida
    parcial: ni ingestion_report.json (produces[0].path(ctx)) ni copia
    alguna en ctx.bronze_dir (que ya existe vacio por el scaffold de
    create_client, DS-ING-1: contrato de errores, solo requires ausente
    aborta; las inconsistencias de datos nunca abortan)."""
    ctx = _build_ctx_fixture_completo(tmp_path)
    mapa_path = ctx.outputs_dir / "020_onboarding/map_client_data.json"
    mapa_path.unlink()

    with pytest.raises(FlowContractError):
        Ingestion().run(ctx)

    ruta_reporte = ctx.outputs_dir / "030_ingestion/ingestion_report.json"
    assert not ruta_reporte.exists()
    assert list(ctx.bronze_dir.iterdir()) == []
